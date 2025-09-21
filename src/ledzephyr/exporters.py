"""
Data export functionality for ledzephyr.

Supports multiple export formats including Excel, PDF, HTML,
and various data visualization options.
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from jinja2 import Template
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ledzephyr.models import ProjectMetrics, TeamMetrics


class ExcelExporter:
    """Export data to Excel format with formatting and charts."""

    def __init__(self):
        self.workbook = None
        self.styles = self._create_styles()

    def _create_styles(self) -> dict[str, Any]:
        """Create reusable Excel styles."""
        return {
            "header": {
                "font": Font(bold=True, color="FFFFFF", size=12),
                "fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                ),
            },
            "subheader": {
                "font": Font(bold=True, size=11),
                "fill": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
                "alignment": Alignment(horizontal="left", vertical="center"),
            },
            "data": {
                "alignment": Alignment(horizontal="left", vertical="center"),
                "border": Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                ),
            },
            "percentage": {
                "alignment": Alignment(horizontal="right", vertical="center"),
                "number_format": "0.00%",
            },
            "number": {
                "alignment": Alignment(horizontal="right", vertical="center"),
                "number_format": "#,##0",
            },
        }

    def export_metrics(
        self,
        metrics_data: dict[str, ProjectMetrics],
        output_path: str | Path,
        include_charts: bool = True,
    ) -> Path:
        """Export metrics to Excel with multiple sheets."""
        output_path = Path(output_path)
        self.workbook = Workbook()

        # Remove default sheet
        self.workbook.remove(self.workbook.active)

        # Create summary sheet
        self._create_summary_sheet(metrics_data)

        # Create detailed sheets for each time window
        for window, metrics in metrics_data.items():
            self._create_metrics_sheet(window, metrics, include_charts)

        # Create comparison sheet
        if len(metrics_data) > 1:
            self._create_comparison_sheet(metrics_data, include_charts)

        # Save workbook
        self.workbook.save(output_path)
        return output_path

    def _create_summary_sheet(self, metrics_data: dict[str, ProjectMetrics]):
        """Create summary sheet with key metrics."""
        ws = self.workbook.create_sheet("Summary")

        # Title
        ws["A1"] = "Ledzephyr Migration Metrics Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:F1")

        # Timestamp
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(italic=True)

        # Headers
        headers = ["Time Window", "Total Tests", "Zephyr", "qTest", "Adoption %", "Active Users"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            for style_key, style_value in self.styles["header"].items():
                setattr(cell, style_key, style_value)

        # Data rows
        row = 5
        for window, metrics in metrics_data.items():
            ws.cell(row=row, column=1, value=window)
            ws.cell(row=row, column=2, value=metrics.total_tests)
            ws.cell(row=row, column=3, value=metrics.zephyr_tests)
            ws.cell(row=row, column=4, value=metrics.qtest_tests)
            ws.cell(row=row, column=5, value=metrics.adoption_ratio)
            ws.cell(row=row, column=6, value=metrics.active_users)

            # Apply formatting
            ws.cell(row=row, column=5).number_format = "0.00%"
            row += 1

        # Auto-fit columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    def _create_metrics_sheet(self, window: str, metrics: ProjectMetrics, include_charts: bool):
        """Create detailed metrics sheet for a time window."""
        ws = self.workbook.create_sheet(f"Metrics - {window}")

        # Project metrics
        row = 1
        ws.cell(row=row, column=1, value=f"Project Metrics - {window} Window")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)

        row += 2
        project_data = [
            ["Metric", "Value"],
            ["Project Key", metrics.project_key],
            ["Time Window", metrics.time_window],
            ["Total Tests", metrics.total_tests],
            ["Zephyr Tests", metrics.zephyr_tests],
            ["qTest Tests", metrics.qtest_tests],
            ["Adoption Ratio", metrics.adoption_ratio],
            ["Active Users", metrics.active_users],
            ["Coverage Parity", metrics.coverage_parity],
            ["Defect Link Rate", metrics.defect_link_rate],
        ]

        for data_row in project_data:
            ws.cell(row=row, column=1, value=data_row[0])
            ws.cell(row=row, column=2, value=data_row[1])

            # Apply formatting
            if row == 3:  # Header row
                for col in range(1, 3):
                    cell = ws.cell(row=row, column=col)
                    for style_key, style_value in self.styles["header"].items():
                        setattr(cell, style_key, style_value)
            else:
                if isinstance(data_row[1], float) and data_row[1] <= 1.0:
                    ws.cell(row=row, column=2).number_format = "0.00%"

            row += 1

        # Team metrics
        if metrics.teams:
            row += 2
            ws.cell(row=row, column=1, value="Team Metrics")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)

            row += 2
            team_headers = ["Team", "Total", "Zephyr", "qTest", "Adoption %", "Users"]
            for col, header in enumerate(team_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                for style_key, style_value in self.styles["header"].items():
                    setattr(cell, style_key, style_value)

            row += 1
            for team in metrics.teams:
                ws.cell(row=row, column=1, value=team.team_name)
                ws.cell(row=row, column=2, value=team.total_tests)
                ws.cell(row=row, column=3, value=team.zephyr_tests)
                ws.cell(row=row, column=4, value=team.qtest_tests)
                ws.cell(row=row, column=5, value=team.adoption_ratio)
                ws.cell(row=row, column=6, value=team.active_users)

                ws.cell(row=row, column=5).number_format = "0.00%"
                row += 1

        # Add charts if requested
        if include_charts and metrics.teams:
            self._add_team_chart(ws, metrics.teams, start_row=row + 2)

    def _add_team_chart(self, ws, teams: list[TeamMetrics], start_row: int):
        """Add bar chart for team metrics."""
        # Prepare data for chart
        chart_data = []
        for team in teams[:10]:  # Limit to top 10 teams
            chart_data.append([team.team_name, team.zephyr_tests, team.qtest_tests])

        # Write chart data
        for i, row_data in enumerate(chart_data):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=10 + j, value=value)

        # Create chart
        chart = BarChart()
        chart.title = "Test Distribution by Team"
        chart.y_axis.title = "Number of Tests"
        chart.x_axis.title = "Team"

        # Set data
        data = Reference(
            ws, min_col=11, min_row=start_row, max_row=start_row + len(chart_data) - 1, max_col=12
        )
        categories = Reference(
            ws, min_col=10, min_row=start_row, max_row=start_row + len(chart_data) - 1
        )

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.series[0].title = "Zephyr"
        chart.series[1].title = "qTest"

        # Add chart to worksheet
        ws.add_chart(chart, f"A{start_row}")

    def _create_comparison_sheet(
        self, metrics_data: dict[str, ProjectMetrics], include_charts: bool
    ):
        """Create comparison sheet across time windows."""
        ws = self.workbook.create_sheet("Comparison")

        # Convert to DataFrame for easier manipulation
        comparison_data = []
        for window, metrics in metrics_data.items():
            comparison_data.append(
                {
                    "Window": window,
                    "Total Tests": metrics.total_tests,
                    "Zephyr": metrics.zephyr_tests,
                    "qTest": metrics.qtest_tests,
                    "Adoption %": metrics.adoption_ratio,
                    "Active Users": metrics.active_users,
                }
            )

        df = pd.DataFrame(comparison_data)

        # Write DataFrame to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Apply header formatting
                if r_idx == 1:
                    for style_key, style_value in self.styles["header"].items():
                        setattr(cell, style_key, style_value)

        # Auto-fit columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)


class PDFExporter:
    """Export data to PDF format with charts and tables."""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._add_custom_styles()

    def _add_custom_styles(self):
        """Add custom paragraph styles."""
        self.styles.add(
            ParagraphStyle(
                name="CustomTitle",
                parent=self.styles["Title"],
                fontSize=24,
                textColor=colors.HexColor("#366092"),
                spaceAfter=30,
            )
        )

        self.styles.add(
            ParagraphStyle(
                name="CustomHeading",
                parent=self.styles["Heading1"],
                fontSize=16,
                textColor=colors.HexColor("#366092"),
                spaceAfter=12,
            )
        )

    def export_metrics(
        self,
        metrics_data: dict[str, ProjectMetrics],
        output_path: str | Path,
        include_charts: bool = True,
    ) -> Path:
        """Export metrics to PDF."""
        output_path = Path(output_path)
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )

        # Build story
        story = []

        # Title page
        story.append(Paragraph("Ledzephyr Migration Metrics Report", self.styles["CustomTitle"]))
        story.append(Spacer(1, 12))
        story.append(
            Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", self.styles["Normal"]
            )
        )
        story.append(Spacer(1, 24))

        # Executive summary
        story.append(Paragraph("Executive Summary", self.styles["CustomHeading"]))
        story.append(self._create_summary_table(metrics_data))
        story.append(Spacer(1, 24))

        # Detailed metrics for each window
        for window, metrics in metrics_data.items():
            story.append(PageBreak())
            story.append(Paragraph(f"Metrics - {window} Window", self.styles["CustomHeading"]))
            story.append(Spacer(1, 12))

            # Project metrics table
            story.append(self._create_metrics_table(metrics))
            story.append(Spacer(1, 24))

            # Team metrics
            if metrics.teams:
                story.append(Paragraph("Team Breakdown", self.styles["Heading2"]))
                story.append(self._create_team_table(metrics.teams))

            # Add charts if requested
            if include_charts:
                story.append(Spacer(1, 24))
                chart_drawing = self._create_chart(metrics)
                if chart_drawing:
                    story.append(chart_drawing)

        # Build PDF
        doc.build(story)
        return output_path

    def _create_summary_table(self, metrics_data: dict[str, ProjectMetrics]) -> Table:
        """Create summary table."""
        data = [["Time Window", "Total Tests", "Zephyr", "qTest", "Adoption %"]]

        for window, metrics in metrics_data.items():
            data.append(
                [
                    window,
                    str(metrics.total_tests),
                    str(metrics.zephyr_tests),
                    str(metrics.qtest_tests),
                    f"{metrics.adoption_ratio:.1%}",
                ]
            )

        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return table

    def _create_metrics_table(self, metrics: ProjectMetrics) -> Table:
        """Create detailed metrics table."""
        data = [
            ["Metric", "Value"],
            ["Project Key", metrics.project_key],
            ["Total Tests", str(metrics.total_tests)],
            ["Zephyr Tests", str(metrics.zephyr_tests)],
            ["qTest Tests", str(metrics.qtest_tests)],
            ["Adoption Ratio", f"{metrics.adoption_ratio:.1%}"],
            ["Active Users", str(metrics.active_users)],
            ["Coverage Parity", f"{metrics.coverage_parity:.1%}"],
            ["Defect Link Rate", f"{metrics.defect_link_rate:.1%}"],
        ]

        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return table

    def _create_team_table(self, teams: list[TeamMetrics]) -> Table:
        """Create team metrics table."""
        data = [["Team", "Total", "Zephyr", "qTest", "Adoption %"]]

        for team in teams[:10]:  # Limit to top 10
            data.append(
                [
                    team.team_name,
                    str(team.total_tests),
                    str(team.zephyr_tests),
                    str(team.qtest_tests),
                    f"{team.adoption_ratio:.1%}",
                ]
            )

        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return table

    def _create_chart(self, metrics: ProjectMetrics) -> Drawing | None:
        """Create a chart drawing."""
        if not metrics.teams:
            return None

        drawing = Drawing(400, 200)

        # Create pie chart for test distribution
        pie = Pie()
        pie.x = 50
        pie.y = 50
        pie.width = 100
        pie.height = 100

        pie.data = [metrics.zephyr_tests, metrics.qtest_tests]
        pie.labels = ["Zephyr", "qTest"]
        pie.slices.strokeWidth = 0.5
        pie.slices[0].fillColor = colors.blue
        pie.slices[1].fillColor = colors.green

        drawing.add(pie)
        return drawing


class HTMLExporter:
    """Export data to HTML format with interactive charts."""

    def __init__(self):
        self.template = self._load_template()

    def _load_template(self) -> Template:
        """Load HTML template."""
        template_str = """
<!DOCTYPE html>
<html>
<head>
    <title>Ledzephyr Metrics Report</title>
    <meta charset="utf-8">
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1 { color: #366092; }
        h2 { color: #4a7ba7; }
        table { border-collapse: collapse; width: 100%; margin: 20px 0; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #366092; color: white; }
        tr:nth-child(even) { background-color: #f2f2f2; }
        .metric-card {
            display: inline-block;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 15px;
            margin: 10px;
            background: #f9f9f9;
        }
        .metric-value { font-size: 24px; font-weight: bold; color: #366092; }
        .metric-label { color: #666; }
        .chart-container { width: 100%; height: 400px; margin: 20px 0; }
    </style>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
</head>
<body>
    <h1>Ledzephyr Migration Metrics Report</h1>
    <p>Generated: {{ timestamp }}</p>

    <h2>Summary</h2>
    <div id="summary-cards">
        {% for window, metrics in metrics_data.items() %}
        <div class="metric-card">
            <div class="metric-label">{{ window }}</div>
            <div class="metric-value">{{ metrics.adoption_ratio|percentage }}</div>
            <div class="metric-label">Adoption Rate</div>
        </div>
        {% endfor %}
    </div>

    <h2>Detailed Metrics</h2>
    <table>
        <thead>
            <tr>
                <th>Time Window</th>
                <th>Total Tests</th>
                <th>Zephyr</th>
                <th>qTest</th>
                <th>Adoption %</th>
                <th>Active Users</th>
            </tr>
        </thead>
        <tbody>
            {% for window, metrics in metrics_data.items() %}
            <tr>
                <td>{{ window }}</td>
                <td>{{ metrics.total_tests }}</td>
                <td>{{ metrics.zephyr_tests }}</td>
                <td>{{ metrics.qtest_tests }}</td>
                <td>{{ metrics.adoption_ratio|percentage }}</td>
                <td>{{ metrics.active_users }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>

    <h2>Charts</h2>
    <div id="adoption-chart" class="chart-container"></div>
    <div id="distribution-chart" class="chart-container"></div>

    <script>
        // Adoption trend chart
        var adoptionData = [{
            x: [{% for window in metrics_data.keys() %}'{{ window }}',{% endfor %}],
            y: [{% for metrics in metrics_data.values() %}{{ metrics.adoption_ratio }},{% endfor %}],
            type: 'scatter',
            mode: 'lines+markers',
            name: 'Adoption Rate'
        }];

        var adoptionLayout = {
            title: 'Adoption Rate Trend',
            xaxis: { title: 'Time Window' },
            yaxis: { title: 'Adoption Rate', tickformat: '.0%' }
        };

        Plotly.newPlot('adoption-chart', adoptionData, adoptionLayout);

        // Test distribution chart
        var distributionData = [
            {% for window, metrics in metrics_data.items() %}
            {
                x: ['Zephyr', 'qTest'],
                y: [{{ metrics.zephyr_tests }}, {{ metrics.qtest_tests }}],
                name: '{{ window }}',
                type: 'bar'
            },
            {% endfor %}
        ];

        var distributionLayout = {
            title: 'Test Distribution',
            xaxis: { title: 'System' },
            yaxis: { title: 'Number of Tests' },
            barmode: 'group'
        };

        Plotly.newPlot('distribution-chart', distributionData, distributionLayout);
    </script>
</body>
</html>
"""
        return Template(template_str)

    def export_metrics(
        self,
        metrics_data: dict[str, ProjectMetrics],
        output_path: str | Path,
    ) -> Path:
        """Export metrics to HTML."""
        output_path = Path(output_path)

        # Prepare template context
        context = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "metrics_data": {
                window: metrics.model_dump() for window, metrics in metrics_data.items()
            },
        }

        # Custom filter for percentage formatting
        def percentage_filter(value):
            return f"{value:.1%}"

        # Render template
        html_content = self.template.render(**context)

        # Write to file
        output_path.write_text(html_content)
        return output_path


class DataExporter:
    """Main data export manager supporting multiple formats."""

    def __init__(self):
        self.excel_exporter = ExcelExporter()
        self.pdf_exporter = PDFExporter()
        self.html_exporter = HTMLExporter()

    def export(
        self,
        metrics_data: dict[str, ProjectMetrics],
        output_path: str | Path,
        output_format: str = "excel",
        include_charts: bool = True,
    ) -> Path:
        """
        Export metrics data to specified format.

        Args:
            metrics_data: Dictionary of metrics by time window
            output_path: Output file path
            output_format: Export format (excel, pdf, html, csv, json)
            include_charts: Include charts in export

        Returns:
            Path to exported file
        """
        output_path = Path(output_path)

        if output_format.lower() == "excel":
            return self.excel_exporter.export_metrics(metrics_data, output_path, include_charts)
        elif output_format.lower() == "pdf":
            return self.pdf_exporter.export_metrics(metrics_data, output_path, include_charts)
        elif output_format.lower() == "html":
            return self.html_exporter.export_metrics(metrics_data, output_path)
        elif output_format.lower() == "csv":
            return self._export_csv(metrics_data, output_path)
        elif output_format.lower() == "json":
            return self._export_json(metrics_data, output_path)
        else:
            raise ValueError(f"Unsupported export format: {output_format}")

    def _export_csv(self, metrics_data: dict[str, ProjectMetrics], output_path: Path) -> Path:
        """Export to CSV format."""
        # Convert to DataFrame
        rows = []
        for window, metrics in metrics_data.items():
            row = {
                "window": window,
                "total_tests": metrics.total_tests,
                "zephyr_tests": metrics.zephyr_tests,
                "qtest_tests": metrics.qtest_tests,
                "adoption_ratio": metrics.adoption_ratio,
                "active_users": metrics.active_users,
                "coverage_parity": metrics.coverage_parity,
                "defect_link_rate": metrics.defect_link_rate,
            }
            rows.append(row)

        df = pd.DataFrame(rows)
        df.to_csv(output_path, index=False)
        return output_path

    def _export_json(self, metrics_data: dict[str, ProjectMetrics], output_path: Path) -> Path:
        """Export to JSON format."""
        json_data = {
            "timestamp": datetime.now().isoformat(),
            "metrics": {window: metrics.model_dump() for window, metrics in metrics_data.items()},
        }

        with open(output_path, "w") as f:
            json.dump(json_data, f, indent=2, default=str)

        return output_path
